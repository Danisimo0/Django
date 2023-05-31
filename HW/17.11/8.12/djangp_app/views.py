from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook


def export_to_excel(request):
    wb = Workbook()

    ws = wb.active

    # Получение данных, которые надо экспортировать в Excel
    data = ["Example Data 1", "Example Data 2", "Example Data 3"]

    # Запись данных  и
    for index, value in enumerate(data, start=1):
        ws.cell(row=index, column=1, value=value)

    ws.cell(row=1, column=1, value="Data")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

    wb.save(response)

    return response


formatted_number = "{:,}".format(1258111).replace(",", " ")
# 2.	Разделите на «разряды» данные (1258111 => 1 258 111).